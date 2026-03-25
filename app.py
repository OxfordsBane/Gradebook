import streamlit as st
import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import range_boundaries, get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Protection
from openpyxl.formatting.rule import Rule, IconSet, FormatObject, CellIsRule, FormulaRule
from openpyxl.workbook.protection import WorkbookProtection, FileSharing
import re
import io
import zipfile

def get_class_info_from_sheet(sheet):
    students = []
    advisor_name = ""
    start_reading = False
    
    for row in sheet.iter_rows(values_only=True):
        if row[1] == "STUDENT NUMBER":
            start_reading = True
            for cell_val in row:
                if cell_val and isinstance(cell_val, str) and "Advisor" in cell_val:
                    advisor_name = cell_val.split(":")[-1].strip()
                    break
            continue
            
        if start_reading:
            if not row[0] or not str(row[0]).strip().isdigit():
                break
            students.append({
                "index": row[0],
                "number": row[1],
                "name": row[2],
                "surname": row[3]
            })
            
    return students, advisor_name

def get_template_student_rows(wb, sheet_idx, start_row=3):
    ws = wb.worksheets[sheet_idx]
    
    for table in ws.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        if min_row <= start_row <= max_row:
            return max_row - start_row + 1
            
    count = 0
    for r in range(start_row, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val is not None and str(val).strip() != "" and str(val).strip() != "0":
            count += 1
        else:
            break
            
    if count > 0:
        return count
        
    if sheet_idx > 0:
        return get_template_student_rows(wb, 0, start_row)
        
    return 30

def shift_formula_rows(formula_str, threshold_row, offset):
    if not formula_str or not isinstance(formula_str, str) or not formula_str.startswith('='):
        return formula_str
        
    def repl(match):
        prefix = match.group(1)
        row_num = int(match.group(2))
        if row_num >= threshold_row:
            return f"{prefix}{row_num + offset}"
        return match.group(0)
        
    return re.sub(r'(?<![A-Za-z])(\$?[A-Za-z]{1,3}\$?)(\d+)\b', repl, formula_str)

def adjust_template_rows_and_tables(ws, num_students, current_rows):
    start_row = 3
    original_last_student_row = start_row + current_rows - 1
    
    # --- GERÇEK SÜTUN SINIRINI BULMA (HAYALET HÜCRE ENGELLEYİCİ) ---
    actual_max_col = 1
    if list(ws.tables.values()):
        for table in ws.tables.values():
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            if max_col > actual_max_col:
                actual_max_col = max_col
    else:
        for c in range(ws.max_column, 0, -1):
            if ws.cell(row=1, column=c).value is not None or ws.cell(row=2, column=c).value is not None:
                actual_max_col = c
                break
        if actual_max_col == 1:
            actual_max_col = ws.max_column

    original_top_borders = []
    original_bottom_borders = []
    internal_horizontal_borders = []
    default_thin = Side(border_style="thin", color="000000")
    
    for c in range(1, actual_max_col + 1):
        cell_first = ws.cell(row=start_row, column=c)
        cell_last = ws.cell(row=original_last_student_row, column=c)
        
        b_top = cell_first.border.top if cell_first.border else None
        b_bot = cell_last.border.bottom if cell_last.border else None
        
        original_top_borders.append(b_top if b_top and b_top.style else default_thin)
        original_bottom_borders.append(b_bot if b_bot and b_bot.style else default_thin)
        
        if current_rows > 1:
            b_mid = cell_first.border.bottom if cell_first.border else None
            internal_horizontal_borders.append(b_mid if b_mid and b_mid.style else default_thin)
        else:
            internal_horizontal_borders.append(default_thin)
            
    action_row_idx = start_row + (current_rows // 2)
    if action_row_idx <= start_row:
        action_row_idx = start_row + 1
    
    offset = 0
    if num_students > current_rows:
        rows_to_add = num_students - current_rows
        ws.insert_rows(action_row_idx, amount=rows_to_add)
        offset = rows_to_add
        
        for r in range(action_row_idx, action_row_idx + rows_to_add):
            for col in range(1, actual_max_col + 1):
                source_cell = ws.cell(row=start_row, column=col)
                target_cell = ws.cell(row=r, column=col)
                if source_cell.has_style:
                    target_cell._style = source_cell._style

    elif num_students < current_rows:
        rows_to_delete = current_rows - num_students
        ws.delete_rows(action_row_idx, amount=rows_to_delete)
        offset = -rows_to_delete

    last_student_row = start_row + num_students - 1

    if offset != 0:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:
                    cell.value = shift_formula_rows(str(cell.value), action_row_idx, offset)

    for table in ws.tables.values():
        ref = table.ref
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        table_offset = max_row - original_last_student_row
        if table_offset < 0:
            table_offset = 0
        new_table_max_row = last_student_row + table_offset
        table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_table_max_row}"

    for r in range(start_row, last_student_row + 1):
        for c in range(1, actual_max_col + 1):
            target_cell = ws.cell(row=r, column=c)
            b_left = target_cell.border.left if target_cell.border else None
            b_right = target_cell.border.right if target_cell.border else None
            
            if r == start_row and r == last_student_row:
                b_top = original_top_borders[c-1]
                b_bottom = original_bottom_borders[c-1]
            elif r == start_row:
                b_top = original_top_borders[c-1]
                b_bottom = internal_horizontal_borders[c-1]
            elif r == last_student_row:
                b_top = internal_horizontal_borders[c-1]
                b_bottom = original_bottom_borders[c-1]
            else:
                b_top = internal_horizontal_borders[c-1]
                b_bottom = internal_horizontal_borders[c-1]
                
            target_cell.border = Border(left=b_left, right=b_right, top=b_top, bottom=b_bottom)

    for col in range(5, actual_max_col + 1):
        master_cell = ws.cell(row=start_row, column=col)
        if master_cell.data_type != 'f':
            master_cell.value = None

    for r in range(start_row + 1, last_student_row + 1):
        for col in range(1, actual_max_col + 1):
            master_cell = ws.cell(row=start_row, column=col)
            target_cell = ws.cell(row=r, column=col)
            
            if master_cell.data_type == 'f' and master_cell.value:
                try:
                    target_cell.value = Translator(master_cell.value, origin=master_cell.coordinate).translate_formula(target_cell.coordinate)
                except:
                    target_cell.value = master_cell.value
            else:
                if col >= 5:
                    target_cell.value = None

    if hasattr(ws, 'conditional_formatting') and hasattr(ws.conditional_formatting, '_cf_rules'):
        ws.conditional_formatting._cf_rules.clear()
        
    if hasattr(ws, 'extLst'):
        ws.extLst = None

    return last_student_row, actual_max_col

def process_class_template(template_bytes, class_name, students, module_name, advisor_name):
    wb = openpyxl.load_workbook(filename=io.BytesIO(template_bytes))
    
    wb.template = False 
    try:
        if hasattr(wb, 'file_sharing'):
            wb.file_sharing = FileSharing(readOnlyRecommended=False)
        else:
            wb.file_sharing = FileSharing(readOnlyRecommended=False)
    except:
        pass
    
    start_row = 3
    first_sheet_last_row = 3
    level_prefix = class_name.split(".")[0].upper()
    
    passwords = {
        "A1": "Esra",
        "A2": "Ceren",
        "B1": "Anna",
        "B2": "Berk"
    }
    
    black_fill = PatternFill(start_color="FF000000", end_color="FF000000", fill_type="solid")
    white_bold_underline_font = Font(color="FFFFFF", bold=True, underline="single")
    rule_diff_ns = FormulaRule(formula=['ABS($N3-$S3)>6'], stopIfTrue=False, fill=black_fill, font=white_bold_underline_font)
    rule_diff_ty = FormulaRule(formula=['ABS($T3-$Y3)>6'], stopIfTrue=False, fill=black_fill, font=white_bold_underline_font)
    
    blue_fill_x = PatternFill(start_color="FF1B587C", end_color="FF1B587C", fill_type="solid")
    white_bold_font_x = Font(color="FFFFFF", bold=True)
    rule_greater_zero = CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=False, fill=blue_fill_x, font=white_bold_font_x)
    
    for i, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        template_student_rows = get_template_student_rows(wb, i, start_row)
        last_student_row, actual_max_col = adjust_template_rows_and_tables(ws, len(students), template_student_rows)
        
        if i == 0:
            first_sheet_last_row = last_student_row
            
        if i > 0:
            if level_prefix == "B1" and sheet_name.lower() in ["midterm", "met"]:
                thick_side = Side(border_style="medium", color="000000")
                thin_side = Side(border_style="thin", color="000000")
                
                thick_cols = []
                if sheet_name.lower() == "midterm":
                    thick_cols = [5, 9, 14, 19, 24] 
                elif sheet_name.lower() == "met":
                    thick_cols = [5, 9, 15, 20, 25, 30] 
                    
                for r in range(start_row, last_student_row + 1):
                    for c in range(5, actual_max_col + 1):
                        target_cell = ws.cell(row=r, column=c)
                        current_b = target_cell.border
                        
                        b_top = current_b.top if current_b and current_b.top and current_b.top.style else thin_side
                        b_bottom = current_b.bottom if current_b and current_b.bottom and current_b.bottom.style else thin_side
                        
                        if r > start_row:
                            b_top = thin_side
                        if r < last_student_row:
                            b_bottom = thin_side
                            
                        b_left = thin_side
                        b_right = thin_side
                        
                        if c in thick_cols:
                            b_left = thick_side
                            b_right = thick_side
                            if r == start_row:
                                b_top = thick_side
                            if r == last_student_row:
                                b_bottom = thick_side
                                
                        target_cell.border = Border(top=b_top, bottom=b_bottom, left=b_left, right=b_right)

            cfvo1 = FormatObject(type='num', val=0)   
            cfvo2 = FormatObject(type='num', val=45)  
            cfvo3 = FormatObject(type='num', val=60)  
            cfvo4 = FormatObject(type='num', val=70)  
            cfvo5 = FormatObject(type='num', val=85)  
            
            icon_set = IconSet(iconSet='5Arrows', cfvo=[cfvo1, cfvo2, cfvo3, cfvo4, cfvo5])
            rule = Rule(type='iconSet', iconSet=icon_set)
            ws.conditional_formatting.add(f"E3:E{last_student_row}", rule)
            
            if sheet_name.lower() == "midterm":
                if level_prefix == "B2":
                    cfvo_i = [FormatObject(type='num', val=v) for v in [0, 6, 12, 18, 24]]
                else:
                    cfvo_i = [FormatObject(type='num', val=v) for v in [0, 4, 8, 12, 16]]
                icon_set_mid = IconSet(iconSet='5Arrows', cfvo=cfvo_i)
                ws.conditional_formatting.add(f"I3:I{last_student_row}", Rule(type='iconSet', iconSet=icon_set_mid))
                
                cfvo_ns = [FormatObject(type='num', val=v) for v in [0, 8, 16, 24, 32]]
                icon_set_ns = IconSet(iconSet='5Arrows', cfvo=cfvo_ns)
                rule_arrows_ns = Rule(type='iconSet', iconSet=icon_set_ns)
                
                ws.conditional_formatting.add(f"N3:N{last_student_row}", rule_diff_ns)
                ws.conditional_formatting.add(f"S3:S{last_student_row}", rule_diff_ns)
                ws.conditional_formatting.add(f"N3:N{last_student_row}", rule_arrows_ns)
                ws.conditional_formatting.add(f"S3:S{last_student_row}", rule_arrows_ns)
                
                ws.conditional_formatting.add(f"X3:X{last_student_row}", rule_greater_zero)

            elif sheet_name.lower() == "met":
                cfvo_met = [FormatObject(type='num', val=v) for v in [0, 8, 16, 24, 32]]
                icon_set_met = IconSet(iconSet='5Arrows', cfvo=cfvo_met)
                rule_arrows_met = Rule(type='iconSet', iconSet=icon_set_met)
                
                if level_prefix == "A1":
                    ws.conditional_formatting.add(f"N3:N{last_student_row}", rule_diff_ns)
                    ws.conditional_formatting.add(f"S3:S{last_student_row}", rule_diff_ns)
                    ws.conditional_formatting.add(f"N3:N{last_student_row}", rule_arrows_met)
                    ws.conditional_formatting.add(f"S3:S{last_student_row}", rule_arrows_met)
                    ws.conditional_formatting.add(f"X3:X{last_student_row}", rule_greater_zero)
                    
                elif level_prefix in ["A2", "B1", "B2"]:
                    ws.conditional_formatting.add(f"T3:T{last_student_row}", rule_diff_ty)
                    ws.conditional_formatting.add(f"Y3:Y{last_student_row}", rule_diff_ty)
                    ws.conditional_formatting.add(f"T3:T{last_student_row}", rule_arrows_met)
                    ws.conditional_formatting.add(f"Y3:Y{last_student_row}", rule_arrows_met)
                    ws.conditional_formatting.add(f"AD3:AD{last_student_row}", rule_greater_zero)
                    
                    if level_prefix == "A2":
                        cfvo_io = [FormatObject(type='num', val=v) for v in [0, 3, 6, 9, 12]]
                    else:
                        cfvo_io = [FormatObject(type='num', val=v) for v in [0, 4, 8, 12, 16]]
                    icon_set_io = IconSet(iconSet='5Arrows', cfvo=cfvo_io)
                    rule_io = Rule(type='iconSet', iconSet=icon_set_io)
                    ws.conditional_formatting.add(f"I3:I{last_student_row}", rule_io)
                    ws.conditional_formatting.add(f"O3:O{last_student_row}", rule_io)
        
    first_sheet = wb.worksheets[0]
    first_sheet.title = class_name
    
    first_sheet["A1"] = f"{class_name} - {module_name}"
    current_font = first_sheet["A1"].font
    if current_font:
        first_sheet["A1"].font = Font(name=current_font.name, size=20, bold=current_font.bold, italic=current_font.italic, color=current_font.color)
    else:
        first_sheet["A1"].font = Font(size=20, bold=True)
        
    first_sheet_name = first_sheet.title
    for i in range(1, len(wb.worksheets)):
        wb.worksheets[i]["A1"].value = f"='{first_sheet_name}'!A1"
    
    advisor_found = False
    for row in first_sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "Advisor" in cell.value:
                cell.value = f"Advisor: {advisor_name}"
                advisor_found = True
                break
        if advisor_found:
            break
    
    for i, student in enumerate(students):
        first_sheet.cell(row=start_row + i, column=1, value=student["index"])
        first_sheet.cell(row=start_row + i, column=2, value=student["number"])
        first_sheet.cell(row=start_row + i, column=3, value=student["name"])
        first_sheet.cell(row=start_row + i, column=4, value=student["surname"])
        
    cfvo_main = [FormatObject(type='num', val=v) for v in [0, 45, 60, 70, 85]]
    icon_set_main = IconSet(iconSet='5Arrows', cfvo=cfvo_main)
    rule_arrows_main = Rule(type='iconSet', iconSet=icon_set_main)
    
    first_sheet.conditional_formatting.add(f"E3:E{first_sheet_last_row}", rule_arrows_main)
    first_sheet.conditional_formatting.add(f"F3:F{first_sheet_last_row}", rule_arrows_main)
    first_sheet.conditional_formatting.add(f"M3:M{first_sheet_last_row}", rule_arrows_main)

    cfvo_L = [FormatObject(type='num', val=v) for v in [0, 46.99, 49.99]]
    first_sheet.conditional_formatting.add(f"L3:L{first_sheet_last_row}", Rule(type='iconSet', iconSet=IconSet(iconSet='3TrafficLights1', cfvo=cfvo_L)))
    
    cfvo_N = [FormatObject(type='num', val=v) for v in [0, 56.99, 59.5]]
    first_sheet.conditional_formatting.add(f"N3:N{first_sheet_last_row}", Rule(type='iconSet', iconSet=IconSet(iconSet='3Symbols2', cfvo=cfvo_N)))
        
    white_bold = Font(color="FFFFFF", bold=True)
    grades = {"F": "FFCC0000", "C": "FF4E8542", "B": "FF1B587C", "A": "FFFFCC00"}
    for grade, color in grades.items():
        first_sheet.conditional_formatting.add(f"O3:O{first_sheet_last_row}", CellIsRule(operator='equal', formula=[f'"{grade}"'], stopIfTrue=True, fill=PatternFill(start_color=color, end_color=color, fill_type="solid"), font=white_bold))
        
    pwd = passwords.get(level_prefix, "1234")
    for ws_to_protect in wb.worksheets:
        ws_to_protect.protection.sheet = True
        ws_to_protect.protection.set_password(pwd)
        
    wb.security = WorkbookProtection(lockStructure=True)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()

st.title("Excel Gradebook Generator")

class_lists_file = st.file_uploader("Class Lists (Excel)", type=["xlsx"])
module_name = st.text_input("Module Name (e.g., Module 3)", value="Module 3")

st.subheader("Gradebook Templates")
col1, col2 = st.columns(2)
with col1:
    a1_template = st.file_uploader("A1 Gradebook", type=["xltx", "xlsx"])
    a2_template = st.file_uploader("A2 Gradebook", type=["xltx", "xlsx"])
with col2:
    b1_template = st.file_uploader("B1 Gradebook", type=["xltx", "xlsx"])
    b2_template = st.file_uploader("B2 Gradebook", type=["xltx", "xlsx"])

if st.button("Generate Gradebooks"):
    templates = {
        "A1": a1_template,
        "A2": a2_template,
        "B1": b1_template,
        "B2": b2_template
    }
    
    if not class_lists_file:
        st.error("Lütfen Class Lists dosyasını yükleyin.")
    else:
        with st.spinner("Dosyalar oluşturuluyor..."):
            class_wb = openpyxl.load_workbook(class_lists_file, data_only=True)
            
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                for sheet_name in class_wb.sheetnames:
                    level = sheet_name.split(".")[0]
                    
                    if level in templates and templates[level]:
                        ws = class_wb[sheet_name]
                        students, advisor_name = get_class_info_from_sheet(ws)
                        
                        if not students:
                            continue
                            
                        file_data = process_class_template(templates[level].getvalue(), sheet_name, students, module_name, advisor_name)
                        zip_file.writestr(f"{level}/{sheet_name} Gradebook.xlsx", file_data)

            zip_buffer.seek(0)
            st.success("Tüm Gradebook dosyaları başarıyla oluşturuldu!")
            st.download_button(
                label="Oluşturulan Dosyaları İndir (ZIP)",
                data=zip_buffer,
                file_name="Gradebooks.zip",
                mime="application/zip"
            )
